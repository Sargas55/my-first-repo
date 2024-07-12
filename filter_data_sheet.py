import openpyxl
import requests
from datetime import datetime

# 本日の日付を取得
today = datetime.now().date()

# APIからシート名を取得する関数
def get_sheet_name_from_api():
    # APIエンドポイント
    api_url = "https://sample-api.com"
    # クエリパラメータ
    params = {
        'key': '1111',
        'date': today.strftime('%Y-%m-%d')  # 本日の日付をYYYY-MM-DD形式で設定
    }
    # APIリクエスト実行
    response = requests.get(api_url, params=params)
    if response.status_code == 200:
        # レスポンスが成功した場合、JSONからファイル名を取得
        return response.json()['result']['data']['fileName']
    else:
        raise Exception("APIからシート名を取得できませんでした。")

# "Book1.xlsx"を開く
try:
    workbook = openpyxl.load_workbook('Book1.xlsx')
except FileNotFoundError:
    print("ファイル 'Book1.xlsx' が見つかりません。")
    exit()

# "Sheet1" を取得
sheet1 = workbook['Sheet1']

# APIからシート名を取得
new_sheet_name = get_sheet_name_from_api()

# 新しいシート名が既に存在するか確認
if new_sheet_name in workbook.sheetnames:
    print(f"エラー: '{new_sheet_name}' シートが既に存在します。")
    exit()

# 新しいシートを作成
sheet2 = workbook.create_sheet(new_sheet_name)

# "Sheet1"のA列の値が「済」ではなく、B列の日付が本日以前の行だけを新しいシートにコピー
for row in sheet1.iter_rows(min_row=2, values_only=True):
    if row[0] != '済' and row[1] is not None and row[1] <= today:
        sheet2.append(row)

# ファイルを保存
workbook.save('Book1.xlsx')

print(f"'{new_sheet_name}'シートにデータをコピーしました。")